# coding=utf-8
import openpyxl
from openpyxl import Workbook, load_workbook
from scrapy import cmdline

from tscrapy_utils.t_global_data import bili_show_list

filename = "obj.xlsx"
wb = load_workbook(filename)
sheet = wb.get_sheet_by_name('Sheet')
start_row = 1
end_row = 1  # len(bili_show_list)
start_column = 1
end_column = start_column
curr_index = 0  # 当前合并的行号
start_index = 0  # 开始合并的行号
last_title = ""
print "=======合并单元格======="
for row in sheet.rows:
    # for cell in row:
    #     print(cell.coordinate, cell.value)
    if 0 == curr_index:
        curr_index += 1
        continue
    curr_index += 1
    title = row[0].value
    if title == last_title:
        for index in range(len(bili_show_list)):
            sheet.merge_cells(start_row=start_index, start_column=index + 1,
                              end_row=curr_index, end_column=index + 1)
    else:
        last_title = title
        start_index = curr_index

wb.save(filename)  # 保存xlsx文件
